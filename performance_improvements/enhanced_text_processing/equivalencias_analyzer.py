#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Enhanced Equivalencias Analysis Module
Data-driven discovery and quality analysis of equivalencias from the database

Author: Augment Agent
Date: 2025-07-25
"""

import sqlite3
import pandas as pd
import json
import os
from collections import defaultdict, Counter
from typing import Dict, List, Tuple, Set
import re

class EquivalenciasAnalyzer:
    """
    Analyze database to discover new equivalencias and improve existing ones
    """
    
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.discovered_equivalencias = {}
        self.quality_scores = {}
        
    def discover_new_equivalencias(self, min_frequency=5, min_sku_overlap=0.7):
        """
        Analyze database to find potential new equivalencias based on SKU patterns
        
        Args:
            min_frequency: Minimum times a description must appear
            min_sku_overlap: Minimum SKU overlap ratio to consider synonyms
        """
        print(f"🔍 Discovering new equivalencias from database...")
        
        if not os.path.exists(self.db_path):
            print(f"❌ Database not found: {self.db_path}")
            return {}
        
        conn = sqlite3.connect(self.db_path)
        
        # Get descriptions that map to same SKUs frequently
        query = """
        SELECT normalized_descripcion, sku, COUNT(*) as frequency
        FROM processed_consolidado 
        WHERE sku IS NOT NULL AND sku != '' AND normalized_descripcion IS NOT NULL
        GROUP BY normalized_descripcion, sku
        HAVING frequency >= ?
        ORDER BY referencia, frequency DESC
        """
        
        df = pd.read_sql(query, conn, params=[min_frequency])
        conn.close()
        
        print(f"  📊 Found {len(df)} description-SKU combinations with frequency >= {min_frequency}")
        
        # Group by SKU to find potential synonyms
        sku_groups = df.groupby('referencia')
        potential_equivalencias = {}
        
        for sku, group in sku_groups:
            descriptions = group['normalized_descripcion'].tolist()
            
            if len(descriptions) > 1:
                # Calculate similarity between descriptions
                similar_groups = self._find_similar_descriptions(descriptions, min_sku_overlap)
                
                for group_id, desc_group in enumerate(similar_groups):
                    if len(desc_group) > 1:
                        # Calculate total frequency for this group
                        total_freq = group[group['normalized_descripcion'].isin(desc_group)]['frequency'].sum()
                        
                        equivalencia_key = f"{sku}_{group_id}"
                        potential_equivalencias[equivalencia_key] = {
                            'referencia': sku,
                            'descriptions': desc_group,
                            'total_frequency': total_freq,
                            'description_count': len(desc_group),
                            'quality_score': self._calculate_equivalencia_quality(desc_group)
                        }
        
        # Sort by quality score
        sorted_equivalencias = dict(sorted(
            potential_equivalencias.items(), 
            key=lambda x: x[1]['quality_score'], 
            reverse=True
        ))
        
        self.discovered_equivalencias = sorted_equivalencias
        
        print(f"  🎯 Discovered {len(sorted_equivalencias)} potential equivalencia groups")
        
        # Show top discoveries
        print(f"\n📋 Top 10 Discovered Equivalencias:")
        for i, (key, data) in enumerate(list(sorted_equivalencias.items())[:10]):
            print(f"  {i+1}. SKU: {data['referencia']} (Quality: {data['quality_score']:.2f})")
            print(f"     Descriptions: {', '.join(data['descriptions'][:3])}{'...' if len(data['descriptions']) > 3 else ''}")
            print(f"     Frequency: {data['total_frequency']}, Count: {data['description_count']}")
        
        return sorted_equivalencias
    
    def _find_similar_descriptions(self, descriptions: List[str], min_overlap: float) -> List[List[str]]:
        """Find groups of similar descriptions based on word overlap"""
        groups = []
        used_descriptions = set()
        
        for i, desc1 in enumerate(descriptions):
            if desc1 in used_descriptions:
                continue
                
            current_group = [desc1]
            used_descriptions.add(desc1)
            
            words1 = set(desc1.lower().split())
            
            for j, desc2 in enumerate(descriptions[i+1:], i+1):
                if desc2 in used_descriptions:
                    continue
                
                words2 = set(desc2.lower().split())
                
                # Calculate Jaccard similarity
                intersection = len(words1.intersection(words2))
                union = len(words1.union(words2))
                
                if union > 0:
                    similarity = intersection / union
                    
                    if similarity >= min_overlap:
                        current_group.append(desc2)
                        used_descriptions.add(desc2)
            
            if len(current_group) > 1:
                groups.append(current_group)
        
        return groups
    
    def _calculate_equivalencia_quality(self, descriptions: List[str]) -> float:
        """Calculate quality score for an equivalencia group"""
        if len(descriptions) < 2:
            return 0.0
        
        # Factors for quality:
        # 1. Number of descriptions (more = better, up to a point)
        # 2. Length similarity (similar lengths = better)
        # 3. Word overlap (more overlap = better)
        
        # Factor 1: Description count (diminishing returns)
        count_score = min(len(descriptions) / 5.0, 1.0)
        
        # Factor 2: Length similarity
        lengths = [len(desc.split()) for desc in descriptions]
        avg_length = sum(lengths) / len(lengths)
        length_variance = sum((l - avg_length) ** 2 for l in lengths) / len(lengths)
        length_score = max(0, 1.0 - (length_variance / 10.0))  # Normalize variance
        
        # Factor 3: Word overlap
        all_words = []
        for desc in descriptions:
            all_words.extend(desc.lower().split())
        
        word_counts = Counter(all_words)
        common_words = sum(1 for count in word_counts.values() if count > 1)
        total_unique_words = len(word_counts)
        
        overlap_score = common_words / total_unique_words if total_unique_words > 0 else 0
        
        # Combined score
        quality_score = (count_score * 0.3 + length_score * 0.3 + overlap_score * 0.4)
        
        return quality_score
    
    def analyze_existing_equivalencias(self, equivalencias_dict: Dict) -> Dict:
        """
        Analyze quality of existing equivalencias based on database data
        """
        print(f"📊 Analyzing quality of existing equivalencias...")
        
        if not os.path.exists(self.db_path):
            print(f"❌ Database not found: {self.db_path}")
            return {}
        
        conn = sqlite3.connect(self.db_path)
        quality_results = {}
        
        for equiv_id, synonyms in equivalencias_dict.items():
            if not isinstance(synonyms, list) or len(synonyms) < 2:
                continue
            
            # Check how often these synonyms map to the same SKUs
            sku_overlap_score = self._calculate_sku_overlap(conn, synonyms)
            
            # Check frequency balance (are all synonyms actually used?)
            frequency_balance = self._calculate_frequency_balance(conn, synonyms)
            
            # Check if synonyms are actually automotive terms
            automotive_relevance = self._calculate_automotive_relevance(synonyms)
            
            overall_quality = (sku_overlap_score * 0.5 + 
                             frequency_balance * 0.3 + 
                             automotive_relevance * 0.2)
            
            quality_results[equiv_id] = {
                'synonyms': synonyms,
                'sku_overlap_score': sku_overlap_score,
                'frequency_balance': frequency_balance,
                'automotive_relevance': automotive_relevance,
                'overall_quality': overall_quality,
                'recommendation': self._get_quality_recommendation(overall_quality)
            }
        
        conn.close()
        
        # Sort by quality
        sorted_results = dict(sorted(
            quality_results.items(), 
            key=lambda x: x[1]['overall_quality'], 
            reverse=True
        ))
        
        self.quality_scores = sorted_results
        
        # Print summary
        high_quality = sum(1 for r in quality_results.values() if r['overall_quality'] >= 0.7)
        medium_quality = sum(1 for r in quality_results.values() if 0.4 <= r['overall_quality'] < 0.7)
        low_quality = sum(1 for r in quality_results.values() if r['overall_quality'] < 0.4)
        
        print(f"  📈 Quality Analysis Results:")
        print(f"     High Quality (≥0.7): {high_quality}")
        print(f"     Medium Quality (0.4-0.7): {medium_quality}")
        print(f"     Low Quality (<0.4): {low_quality}")
        
        return sorted_results
    
    def _calculate_sku_overlap(self, conn: sqlite3.Connection, synonyms: List[str]) -> float:
        """Calculate how often synonyms map to the same SKUs"""
        sku_sets = []
        
        for synonym in synonyms:
            query = """
            SELECT DISTINCT sku 
            FROM processed_consolidado 
            WHERE normalized_descripcion LIKE ? AND sku IS NOT NULL AND sku != ''
            """
            
            cursor = conn.cursor()
            cursor.execute(query, [f'%{synonym}%'])
            skus = {row[0] for row in cursor.fetchall()}
            
            if skus:
                sku_sets.append(skus)
        
        if len(sku_sets) < 2:
            return 0.0
        
        # Calculate Jaccard similarity between SKU sets
        intersection = set.intersection(*sku_sets)
        union = set.union(*sku_sets)
        
        return len(intersection) / len(union) if union else 0.0
    
    def _calculate_frequency_balance(self, conn: sqlite3.Connection, synonyms: List[str]) -> float:
        """Calculate how balanced the usage frequency is among synonyms"""
        frequencies = []
        
        for synonym in synonyms:
            query = """
            SELECT COUNT(*) 
            FROM processed_consolidado 
            WHERE normalized_descripcion LIKE ?
            """
            
            cursor = conn.cursor()
            cursor.execute(query, [f'%{synonym}%'])
            freq = cursor.fetchone()[0]
            frequencies.append(freq)
        
        if not frequencies or max(frequencies) == 0:
            return 0.0
        
        # Calculate coefficient of variation (lower = more balanced)
        mean_freq = sum(frequencies) / len(frequencies)
        variance = sum((f - mean_freq) ** 2 for f in frequencies) / len(frequencies)
        std_dev = variance ** 0.5
        
        cv = std_dev / mean_freq if mean_freq > 0 else float('inf')
        
        # Convert to score (lower CV = higher score)
        balance_score = max(0, 1.0 - (cv / 2.0))  # Normalize CV
        
        return balance_score
    
    def _calculate_automotive_relevance(self, synonyms: List[str]) -> float:
        """Calculate how relevant synonyms are to automotive domain"""
        automotive_keywords = {
            'parachoques', 'paragolpes', 'bumper', 'defensa',
            'puerta', 'door', 'portezuela',
            'faro', 'farola', 'luz', 'light', 'headlight',
            'espejo', 'mirror', 'retrovisor',
            'capo', 'hood', 'bonnet',
            'guardafango', 'fender', 'guardabarro',
            'vidrio', 'glass', 'cristal', 'parabrisas',
            'motor', 'engine', 'propulsor',
            'delantero', 'trasero', 'front', 'rear',
            'izquierdo', 'derecho', 'left', 'right',
            'superior', 'inferior', 'upper', 'lower'
        }
        
        relevant_count = 0
        total_words = 0
        
        for synonym in synonyms:
            words = synonym.lower().split()
            total_words += len(words)
            
            for word in words:
                if word in automotive_keywords:
                    relevant_count += 1
        
        return relevant_count / total_words if total_words > 0 else 0.0
    
    def _get_quality_recommendation(self, quality_score: float) -> str:
        """Get recommendation based on quality score"""
        if quality_score >= 0.8:
            return "✅ EXCELLENT - Keep as is"
        elif quality_score >= 0.6:
            return "✅ GOOD - Minor improvements possible"
        elif quality_score >= 0.4:
            return "⚠️ FAIR - Needs review and improvement"
        else:
            return "❌ POOR - Consider removing or major revision"
    
    def generate_equivalencias_report(self) -> str:
        """Generate comprehensive equivalencias analysis report"""
        report = f"""
🔍 EQUIVALENCIAS ANALYSIS REPORT
{'='*50}

📊 DISCOVERED EQUIVALENCIAS:
   Total discovered: {len(self.discovered_equivalencias)}
   
   Top 5 Recommendations:"""
        
        for i, (key, data) in enumerate(list(self.discovered_equivalencias.items())[:5]):
            report += f"""
   {i+1}. SKU: {data['referencia']} (Quality: {data['quality_score']:.2f})
      Descriptions: {', '.join(data['descriptions'])}
      Frequency: {data['total_frequency']}"""
        
        if self.quality_scores:
            report += f"""

📈 EXISTING EQUIVALENCIAS QUALITY:
   Total analyzed: {len(self.quality_scores)}
   
   Quality Distribution:"""
            
            high = sum(1 for r in self.quality_scores.values() if r['overall_quality'] >= 0.7)
            medium = sum(1 for r in self.quality_scores.values() if 0.4 <= r['overall_quality'] < 0.7)
            low = sum(1 for r in self.quality_scores.values() if r['overall_quality'] < 0.4)
            
            report += f"""
   • High Quality (≥0.7): {high}
   • Medium Quality (0.4-0.7): {medium}
   • Low Quality (<0.4): {low}
   
   Lowest Quality Equivalencias (need attention):"""
            
            worst = sorted(self.quality_scores.items(), key=lambda x: x[1]['overall_quality'])[:3]
            for i, (equiv_id, data) in enumerate(worst):
                report += f"""
   {i+1}. {equiv_id}: {data['synonyms']} (Quality: {data['overall_quality']:.2f})
      Recommendation: {data['recommendation']}"""
        
        return report
    
    def export_discoveries_to_excel(self, output_path: str):
        """Export discovered equivalencias to Excel format"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Discovered Equivalencias"
            
            # Headers
            headers = ['referencia', 'Descriptions', 'Total Frequency', 'Description Count', 'Quality Score']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            for row, (key, data) in enumerate(self.discovered_equivalencias.items(), 2):
                ws.cell(row=row, column=1, value=data['referencia'])
                ws.cell(row=row, column=2, value=', '.join(data['descriptions']))
                ws.cell(row=row, column=3, value=data['total_frequency'])
                ws.cell(row=row, column=4, value=data['description_count'])
                ws.cell(row=row, column=5, value=data['quality_score'])
            
            wb.save(output_path)
            print(f"📁 Exported discoveries to: {output_path}")
            
        except ImportError:
            print("⚠️ openpyxl not available - cannot export to Excel")
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")


def analyze_equivalencias_from_database(db_path: str, existing_equivalencias: Dict = None):
    """
    Main function to analyze equivalencias from database
    """
    analyzer = EquivalenciasAnalyzer(db_path)
    
    # Discover new equivalencias
    discoveries = analyzer.discover_new_equivalencias(min_frequency=5, min_sku_overlap=0.6)
    
    # Analyze existing equivalencias if provided
    if existing_equivalencias:
        quality_analysis = analyzer.analyze_existing_equivalencias(existing_equivalencias)
    
    # Generate report
    report = analyzer.generate_equivalencias_report()
    print(report)
    
    return analyzer
